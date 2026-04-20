#!/usr/bin/env python3
"""Generate exercise_full_review.xlsx for comprehensive library review."""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open('assets/data/exercise_library.json') as f:
    lib = json.load(f)

# Smart target_focus suggestion builder
MUSCLE_TO_TF = {
    'Chest': ['Mid Chest'],
    'Chest (inner)': ['Mid Chest'],
    'Upper Chest': ['Upper Chest'],
    'Lower Chest': ['Lower Chest'],
    'Quads': ['Quads'],
    'Quads (eccentric)': ['Quads'],
    'Hamstrings': ['Hamstrings'],
    'Glutes': ['Glutes'],
    'Calves': ['Calves'],
    'Calves (Gastrocnemius)': ['Calves'],
    'Calves (Soleus)': ['Calves'],
    'Core': ['Core'],
    'Shoulders': ['Shoulders'],
    'Front Deltoid': ['Front Delts'],
    'Front Delts': ['Front Delts'],
    'Lateral Delts': ['Lateral Delts'],
    'Rear Delts': ['Rear Delts'],
    'Triceps': ['Triceps'],
    'Biceps': ['Biceps'],
    'Lats': ['Lats'],
    'Lower Back': ['Lower Back'],
    'Traps': ['Traps'],
    'Upper Traps': ['Upper Traps'],
    'Adductors': ['Hip'],
    'Forearms': ['Forearms'],
    'Brachialis': ['Biceps'],
    'Rhomboids': ['Mid Back (Thickness)'],
}


def suggest_tf(e):
    current = e.get('target_focus', '')
    pm = e.get('primary_muscles', [])
    suggested = set()
    if current:
        suggested.add(current)
    for m in pm:
        for t in MUSCLE_TO_TF.get(m, []):
            suggested.add(t)
    return sorted(suggested)


exercises = sorted(
    lib,
    key=lambda e: (
        e.get('movement_pattern', ''),
        e.get('exercise_type', ''),
        e.get('name', ''),
    ),
)

# === Styles ===
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_EDIT_FILL = PatternFill(start_color='F57F17', end_color='F57F17', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SECTION_FONT = Font(name='Calibri', bold=True, size=10, color='1F4E79')
EDIT_FILL = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')
READONLY_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
LIGHT_EDIT_FILL = PatternFill(start_color='FFFFF0', end_color='FFFFF0', fill_type='solid')
RED_FILL = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
NORMAL_FONT = Font(name='Calibri', size=9)
EDIT_FONT = Font(name='Calibri', size=9, bold=True, color='1565C0')
READONLY_FONT = Font(name='Calibri', size=9, color='757575')
thin_border = Border(
    left=Side(style='thin', color='CFD8DC'),
    right=Side(style='thin', color='CFD8DC'),
    top=Side(style='thin', color='CFD8DC'),
    bottom=Side(style='thin', color='CFD8DC'),
)

# === Column definitions ===
# (col_letter, header, width, editable, field_key)
columns = [
    ('A', 'ID', 6, False, 'id'),
    ('B', 'Exercise Name', 30, True, 'name'),
    ('C', 'movement_pattern', 18, True, 'movement_pattern'),
    ('D', 'exercise_type', 14, True, 'exercise_type'),
    ('E', 'target_focus\n(comma-sep)', 28, True, 'target_focus'),
    ('F', 'primary_muscles\n(comma-sep)', 28, True, 'primary_muscles'),
    ('G', 'secondary_muscles\n(comma-sep)', 28, True, 'secondary_muscles'),
    ('H', 'equipment_tier\n(comma-sep)', 28, True, 'equipment_tier'),
    ('I', 'suitable_for\n(comma-sep)', 22, True, 'suitable_for'),
    ('J', 'priority_tier', 8, True, 'priority_tier'),
    ('K', 'rep_range', 10, True, 'rep_range'),
    ('L', 'logging_type', 14, True, 'logging_type'),
    ('M', 'injury_contra\n(comma-sep)', 22, True, 'injury_contraindications'),
    ('N', 'equipment_needed\n(comma-sep)', 25, True, 'equipment_needed'),
    ('O', 'is_foundational', 12, True, 'is_foundational'),
    ('P', 'is_bilateral', 10, True, 'is_bilateral'),
    ('Q', 'cns_demand', 8, True, 'cns_demand'),
    ('R', 'standard_swap', 25, True, 'standard_swap'),
    ('S', 'default_sets', 8, True, 'default_sets'),
    ('T', 'default_reps', 8, True, 'default_reps'),
    ('U', 'default_rest_secs', 10, True, 'default_rest_secs'),
]

wb = Workbook()

# ========== SHEET 1: FULL REVIEW ==========
ws = wb.active
ws.title = 'Exercise Review'

# Headers
for col_letter, title, width, editable, _ in columns:
    cell = ws[f'{col_letter}1']
    cell.value = title
    cell.font = HEADER_FONT
    cell.fill = HEADER_EDIT_FILL if editable else HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='bottom')
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[1].height = 35

# Instructions row
ws.merge_cells('A2:U2')
cell = ws['A2']
cell.value = (
    'ALL orange-header columns are EDITABLE. Array fields use comma-separated values. '
    'Column E (target_focus) pre-filled with smart suggestions. '
    'Grey column (ID) is read-only. Grouped by movement pattern.'
)
cell.font = Font(name='Calibri', size=9, italic=True, color='795548')
cell.alignment = Alignment(wrap_text=True)
ws.row_dimensions[2].height = 28

row = 3
current_mp = None

for ex in exercises:
    mp = ex.get('movement_pattern', '')

    # Section header
    if mp != current_mp:
        current_mp = mp
        count = sum(1 for x in exercises if x.get('movement_pattern') == mp)
        last_col = columns[-1][0]
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws[f'A{row}']
        cell.value = f'{current_mp.upper()} ({count} exercises)'
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        ws.row_dimensions[row].height = 20
        row += 1

    for col_idx, (col_letter, _, _, editable, field_key) in enumerate(columns, 1):
        val = ex.get(field_key, '')

        # target_focus: use suggested array
        if field_key == 'target_focus':
            val = ', '.join(suggest_tf(ex))
        elif isinstance(val, list):
            val = ', '.join(str(v) for v in val)
        elif isinstance(val, bool):
            val = str(val).lower()
        elif val is None:
            val = ''

        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        if not editable:
            cell.font = READONLY_FONT
            cell.fill = READONLY_FILL
        elif field_key == 'target_focus':
            cell.font = EDIT_FONT
            cell.fill = EDIT_FILL
        else:
            cell.font = EDIT_FONT
            cell.fill = LIGHT_EDIT_FILL

    row += 1

ws.freeze_panes = 'C3'
ws.auto_filter.ref = f'A2:U{row-1}'

# ========== SHEET 2: VALID VALUES ==========
ref = wb.create_sheet('Valid Values')

ref_data = [
    ('movement_pattern', [
        'horizontal_push', 'vertical_push', 'horizontal_pull', 'vertical_pull',
        'knee_dominant', 'hip_dominant', 'core', 'elbow_flexion', 'elbow_extension',
        'shoulder_isolation', 'hip_isolation', 'cardio', 'warmup', 'cooldown', 'flexibility',
    ]),
    ('exercise_type', [
        'compound', 'isolation', 'calisthenics', 'cardio', 'flexibility',
        'static_stretch', 'dynamic_stretch', 'activation', 'recovery',
    ]),
    ('equipment_tier', ['bodyweight', 'home_dumbbells', 'basic_gym', 'full_gym']),
    ('suitable_for', ['Beginner', 'Intermediate', 'Advanced']),
    ('target_focus vocab', [
        'Mid Chest', 'Upper Chest', 'Lower Chest',
        'Front Delts', 'Lateral Delts', 'Rear Delts', 'Shoulders',
        'Lats', 'Lats (Width)', 'Lats (Thickness)', 'Mid Back (Thickness)',
        'Quads', 'Hamstrings', 'Glutes', 'Calves', 'Hip',
        'Biceps', 'Biceps (Long Head)', 'Biceps (Short Head)', 'Brachialis', 'Forearms',
        'Triceps', 'Triceps (Long Head)',
        'Core', 'Core/Obliques', 'Lower Back', 'Traps', 'Upper Traps',
        'full body', 'full back',
    ]),
    ('logging_type', [
        'weight_reps', 'bodyweight_reps', 'weighted_bodyweight',
        'timed', 'cardio', 'distance',
    ]),
    ('priority_tier', ['1 = primary compound', '2 = secondary', '3 = accessory']),
    ('injury_contra', [
        'shoulder', 'knee', 'lower_back', 'hip', 'elbow',
        'wrist', 'ankle', 'neck', 'hamstring',
    ]),
    ('rep_range', ['5-8', '8-12', '10-15', '12-15', '12-20', '20-30', '30-60']),
]

col = 1
for header, values in ref_data:
    cell = ref.cell(row=1, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ref.column_dimensions[chr(64 + col)].width = 22
    for i, v in enumerate(values, 2):
        cell = ref.cell(row=i, column=col, value=v)
        cell.font = NORMAL_FONT
    col += 1

ref.freeze_panes = 'A2'

# ========== SHEET 3: PROBLEM SLOTS ==========
slots_ws = wb.create_sheet('Generator Slot Demands')

slot_headers = [
    'Split', 'Day', 'targetMuscle', 'movementPattern',
    'exerciseType', 'priority', 'Library matches?',
]
for i, h in enumerate(slot_headers, 1):
    cell = slots_ws.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    slots_ws.column_dimensions[chr(64 + i)].width = 20

problem_slots = [
    ('5-day build_muscle', 'Push A', 'Lower Chest', 'horizontal_push',
     'isolation', 3, 'NO - 0 hp isolations target Lower Chest'),
    ('5-day build_muscle', 'Push B', 'Lower Chest', 'horizontal_push',
     'isolation', 3, 'NO - 0 hp isolations target Lower Chest'),
    ('5-day build_muscle', 'Legs A', 'Hamstrings', 'knee_dominant',
     'isolation', 2, 'NO - ham isolations are hip_dominant'),
    ('5-day build_muscle', 'Legs B', 'Hamstrings', 'knee_dominant',
     'isolation', 2, 'NO - ham isolations are hip_dominant'),
    ('6-day build_muscle', 'Push', 'Lower Chest', 'horizontal_push',
     'isolation', 2, 'NO - 0 hp isolations target Lower Chest'),
    ('6-day build_muscle', 'Legs', 'Hamstrings', 'knee_dominant',
     'isolation', 2, 'NO - ham isolations are hip_dominant'),
    ('4-day strength', 'Legs', 'Hamstrings', 'knee_dominant',
     'isolation', 2, 'NO - ham isolations are hip_dominant'),
    ('5-day strength', 'Push A', 'Lower Chest', 'horizontal_push',
     'isolation', 3, 'NO - 0 hp isolations target Lower Chest'),
    ('5-day strength', 'Legs', 'Hamstrings', 'knee_dominant',
     'isolation', 2, 'NO - ham isolations are hip_dominant'),
    ('various', 'Shoulders', 'Shoulders', 'vertical_push',
     'compound', 1, 'PARTIAL - only Z Press has tf=Shoulders'),
    ('various', 'Pull', 'Mid Back', 'horizontal_pull',
     'compound', 1, 'PARTIAL - most say Lats (Thickness)'),
]

for i, slot in enumerate(problem_slots, 2):
    for j, val in enumerate(slot, 1):
        cell = slots_ws.cell(row=i, column=j, value=val)
        cell.font = NORMAL_FONT
        cell.border = thin_border
        if 'NO' in str(slot[-1]):
            cell.fill = RED_FILL
        elif 'PARTIAL' in str(slot[-1]):
            cell.fill = ORANGE_FILL

slots_ws.freeze_panes = 'A2'

# ========== SAVE ==========
out_path = 'test/plan_generator/exercise_full_review.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Sheet 1: Exercise Review - {len(exercises)} exercises, {len(columns)} columns')
print(f'Sheet 2: Valid Values - reference for allowed values')
print(f'Sheet 3: Generator Slot Demands - {len(problem_slots)} known problem slots')
