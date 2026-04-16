#!/usr/bin/env python3
"""
Import Gemini-enriched Excel into exercise_library.json.

Reads: assets/Fully_Enriched_Exercise_Library_made_from_gemini.xlsx
Merges with: assets/data/exercise_library.json (preserves 17 JSON-only fields)
Writes: assets/data/exercise_library.json

Multi-value fields (comma-separated in Excel) become JSON arrays:
  target_focus, movement_pattern, exercise_type, primary_muscles,
  secondary_muscles, equipment_tier, suitable_for,
  injury_contraindications, equipment_needed
"""

import json
import os
import sys
from pathlib import Path
from openpyxl import load_workbook

# Resolve paths relative to repo root
REPO_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_PATH = REPO_ROOT / 'assets' / 'Fully_Enriched_Exercise_Library_made_from_gemini.xlsx'
JSON_PATH = REPO_ROOT / 'assets' / 'data' / 'exercise_library.json'

# Excel header → JSON field name mapping
HEADER_MAP = {
    'ID': 'id',
    'Exercise Name': 'name',
    'movement_pattern': 'movement_pattern',
    'exercise_type': 'exercise_type',
    'target_focus': 'target_focus',
    'primary_muscles': 'primary_muscles',
    'secondary_muscles': 'secondary_muscles',
    'equipment_tier': 'equipment_tier',
    'suitable_for': 'suitable_for',
    'priority_tier': 'priority_tier',
    'rep_range': 'rep_range',
    'logging_type': 'logging_type',
    'injury_contra': 'injury_contraindications',
    'equipment_needed': 'equipment_needed',
    'is_foundational': 'is_foundational',
    'is_bilateral': 'is_bilateral',
    'cns_demand': 'cns_demand',
    'standard_swap': 'standard_swap',
    'default_sets': 'default_sets',
    'default_reps': 'default_reps',
    'default_rest_secs': 'default_rest_secs',
}

# Fields that should be parsed as comma-separated arrays
ARRAY_FIELDS = {
    'target_focus', 'movement_pattern', 'exercise_type',
    'primary_muscles', 'secondary_muscles', 'equipment_tier',
    'suitable_for', 'injury_contraindications', 'equipment_needed',
}

# Integer fields
INT_FIELDS = {'priority_tier', 'cns_demand', 'default_sets', 'default_rest_secs'}

# Boolean fields
BOOL_FIELDS = {'is_foundational', 'is_bilateral'}

# JSON-only fields to preserve during merge (not in Excel)
PRESERVE_FIELDS = {
    'coaching_cues', 'breathing_cue', 'common_mistakes',
    'warmup_protocol', 'pro_tip', 'tempo', 'met_value',
    'cal_per_set_est', 'is_indian_context', 'indian_alternative',
    'source', 'image_start_url', 'image_end_url', 'gif_url',
    'category', 'difficulty_level', 'is_active', 'name_aliases',
}


def clean_header(raw: str) -> str:
    """Strip \\n(comma-sep) suffixes from Excel headers."""
    if raw is None:
        return ''
    # Headers like "target_focus\n(comma-sep)" -> "target_focus"
    return raw.split('\n')[0].strip()


def parse_value(val, field_name: str):
    """Convert an Excel cell value to the correct JSON type."""
    # Handle None / empty
    if val is None:
        if field_name in ARRAY_FIELDS:
            return []
        if field_name in INT_FIELDS:
            return 0
        if field_name in BOOL_FIELDS:
            return False
        return ''

    # Convert to string for processing
    s = str(val).strip()

    # Handle "None" string
    if s.lower() == 'none' or s == '':
        if field_name in ARRAY_FIELDS:
            return []
        if field_name in INT_FIELDS:
            return 0
        if field_name in BOOL_FIELDS:
            return False
        return ''

    # Array fields: split on comma
    if field_name in ARRAY_FIELDS:
        parts = [p.strip() for p in s.split(',') if p.strip()]
        return parts

    # Integer fields
    if field_name in INT_FIELDS:
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return 0

    # Boolean fields
    if field_name in BOOL_FIELDS:
        return s.lower() in ('true', '1', 'yes')

    # String fields
    return s


def main():
    # Load existing JSON
    print(f'Loading existing JSON: {JSON_PATH}')
    with open(JSON_PATH, encoding='utf-8') as f:
        existing = json.load(f)

    existing_by_id = {e['id']: e for e in existing}
    existing_by_name = {e['name'].lower(): e for e in existing}
    print(f'  Existing exercises: {len(existing)}')

    # Load Excel
    print(f'Loading Excel: {EXCEL_PATH}')
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    print(f'  Sheet: {ws.title}')

    # Build header map from row 1
    rows = list(ws.iter_rows(values_only=True))
    raw_headers = rows[0]
    headers = []
    for h in raw_headers:
        cleaned = clean_header(str(h) if h else '')
        mapped = HEADER_MAP.get(cleaned, cleaned)
        headers.append(mapped)

    print(f'  Headers: {headers}')
    print(f'  Total rows (incl header): {len(rows)}')

    # Process data rows
    imported = []
    skipped = []

    for row_idx, row in enumerate(rows[1:], start=2):
        # Get ID (first column)
        raw_id = str(row[0]).strip() if row[0] else ''

        # Skip section header rows (merged cells with pattern names like
        # "ELBOW_EXTENSION (9 exercises)"). Valid IDs are E + digits only.
        if not raw_id.startswith('E') or not raw_id[1:].isdigit():
            skipped.append((row_idx, raw_id, row[1] if len(row) > 1 else ''))
            continue

        # Build exercise dict from Excel
        excel_data = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row):
                excel_data[header] = parse_value(row[col_idx], header)
            else:
                excel_data[header] = parse_value(None, header)

        # Merge with existing JSON data
        ex_id = excel_data.get('id', '')
        existing_entry = existing_by_id.get(ex_id)

        if existing_entry is None:
            # Try matching by name
            ex_name = excel_data.get('name', '').lower()
            existing_entry = existing_by_name.get(ex_name)

        if existing_entry:
            # Start with existing, then overlay Excel fields
            merged = dict(existing_entry)
            for key, val in excel_data.items():
                if key in PRESERVE_FIELDS:
                    continue  # Don't overwrite preserved fields
                merged[key] = val
        else:
            # New exercise (no existing match)
            merged = excel_data
            print(f'  WARNING: No existing match for {ex_id} "{excel_data.get("name", "?")}"')

        imported.append(merged)

    print(f'\n  Imported: {len(imported)} exercises')
    print(f'  Skipped rows: {len(skipped)}')
    for row_num, rid, rname in skipped:
        print(f'    Row {row_num}: ID="{rid}" Name="{rname}"')

    # Check for exercises in JSON but not in Excel (besides Malkhamb)
    imported_ids = {e['id'] for e in imported}
    missing = [e for e in existing if e['id'] not in imported_ids]
    if missing:
        print(f'\n  Exercises in JSON but not in Excel ({len(missing)}):')
        for m in missing:
            print(f'    {m["id"]}: {m["name"]}')
        print('  These will be DROPPED from output.')

    # Validation: check all array fields are actually arrays
    issues = 0
    for ex in imported:
        for field in ARRAY_FIELDS:
            val = ex.get(field)
            if val is not None and not isinstance(val, list):
                print(f'  ERROR: {ex["id"]} {ex["name"]}: {field} is not a list: {val}')
                issues += 1

    if issues:
        print(f'\n  {issues} validation issues found!')
        sys.exit(1)

    # Sort by ID for stable output
    imported.sort(key=lambda e: e.get('id', ''))

    # Write output
    print(f'\nWriting {len(imported)} exercises to {JSON_PATH}')
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(imported, f, indent=2, ensure_ascii=False)

    # Summary stats
    multi_mp = sum(1 for e in imported if len(e.get('movement_pattern', [])) > 1)
    multi_et = sum(1 for e in imported if len(e.get('exercise_type', [])) > 1)
    multi_tf = sum(1 for e in imported if len(e.get('target_focus', [])) > 1)
    print(f'\n  Multi-value movement_pattern: {multi_mp}')
    print(f'  Multi-value exercise_type: {multi_et}')
    print(f'  Multi-value target_focus: {multi_tf}')
    print('\nDone!')


if __name__ == '__main__':
    main()
